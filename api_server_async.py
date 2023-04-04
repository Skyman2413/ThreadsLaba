import asyncio
import os
import shutil
from asyncio import Queue, Lock
from datetime import datetime
from pathlib import Path
import random
import aiologger
import openpyxl
from aiohttp import web
from aiologger.levels import LogLevel

logger = aiologger.Logger(name="logg", level=LogLevel.INFO)

queue_to_consume = Queue()
monitoring_list = {}


async def get_file(request):
    pass


async def get_queue_info(request):
    s = request
    pass


# producer
async def post_add_item(request):
    s = request
    pass


# исполнитель, генерирующий файлы (consumer)
async def worker(name: str):
    while True:
        logger.info(f"{name} запущен")
        item = await queue_to_consume.get()
        logger.info(f"Начало работы над элементом с id {item['id']} by {name} "
              f" в {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        item["worker"] = f"Worker{name}"
        item["state"] = "in_progress"
        lock = Lock()
        # блокируем словарь для выполнения не атомарных операций
        with lock:
            monitoring_list[item["id"]]["state"] = "in progress"
            monitoring_list[item["id"]]["worker"] = name
            monitoring_list[item["id"]]["start_date"] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        try:
            # такой код будет работать вне зависимости от ОС
            path_to_excel = Path(os.getcwd(), "files_to_send", f"{item['id']}.xlsx")
            path_to_template = Path(os.getcwd(), "Товарный-чек.xlsx")
            shutil.copyfile(path_to_template, path_to_excel)
            wb = openpyxl.load_workbook(path_to_excel)
            ws = wb.active
            ws["A1"].value = item["org_name"]
            title_cell = ws["A4"].value
            title_cell = title_cell.replace("id", str(item["id"])).replace("date", item["date"])
            ws["A4"].value = title_cell
            all_sum = 0
            i = 6
            for product in item["product_list"]:
                ws[f"A{i}"].value = i - 5
                ws[f"B{i}"].value = product[0]
                ws[f"H{i}"].value = product[1]
                ws[f"J{i}"].value = product[2]
                ws[f"M{i}"].value = product[1] * product[2]
                all_sum += product[1] * product[2]
                i += 1
            ws["M16"].value = all_sum
            wb.save(path_to_excel)
            wb.close()
            await asyncio.sleep(random.randint(5, 25))
            # блокируем словарь для выполнения не атомарных операций
            with lock:
                monitoring_list[item["id"]]["state"] = "Done Successfully"
        except Exception as e:
            logger.info(e)
            # блокируем словарь для выполнения не атомарных операций
            with lock:
                monitoring_list[item["id"]]["state"] = "Error"
        # блокируем словарь для выполнения не атомарных операций
        with lock:
            monitoring_list[item["id"]]["end_date"] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        queue_to_consume.task_done()
        logger.info(f"Конец работы над элементом с id {item['id']} by {name}"
              f" в {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")


async def main():
    app = web.Application()
    app.add_routes([web.get('/get_file', get_file)])
    app.add_routes([web.get('/queue_info', get_queue_info)])
    app.add_routes([web.post('/add_item', post_add_item)])
    '''producers = [asyncio.create_task(worker(f"Worker {i}")) for i in range(3)]
    await asyncio.gather(*producers)'''
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, 'localhost', 8080)
    await site.start()
    while True:
        await asyncio.sleep(3600)



if __name__ == "__main__":
    asyncio.run(main())
