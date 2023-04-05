import asyncio
import logging
import os
import shutil
from threading import Lock, Thread
from asyncio import Queue
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import time
from pathlib import Path
import random
import openpyxl
from aiohttp import web

index = 0
max_id = 0
queue_to_consume = Queue()
monitoring_list = {}
routes = web.RouteTableDef()
loop = None

@routes.get('/get_file')
async def get_file(request):
    try:
        s = await request.json()
        id = s["id"]
        if not os.path.exists(Path(os.getcwd(), "files_to_send", f"{id}.xlsx")):
            return web.HTTPFound
        return web.FileResponse(path=Path(os.getcwd(), "files_to_send", f"{id}.xlsx"), status=200)
    except Exception as e:
        logging.info(e)
        return web.HTTPBadRequest


@routes.get('/queue_info')
async def get_queue_info(request):
    return web.json_response(monitoring_list)


# producer
@routes.post('/add_item')
async def post_add_item(request):
    global max_id
    s = await request.json()
    status_code = 200
    try:
        max_id = max_id + 1
        id = max_id
        new_element = {"org_name": s["org_name"], "date": s["date"],
                       "product_list": s["product_list"], "state": "new", "worker": "",
                       "created_at":datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
                       "id": id}
        logging.info("Добавление элемента в очередь" + str(new_element))
        await queue_to_consume.put(new_element)
        response_body = {"id": new_element["id"]}
        lock = asyncio.Lock()
        # блокируем словарь для выполнения не атомарных операций
        await lock.acquire()
        monitoring_list[id] = new_element
        lock.release()
        await asyncio.sleep(2)
    except Exception as e:
        logging.info(e)
        return web.HTTPBadRequest
    return web.json_response(status=status_code, data=response_body)


# исполнитель, генерирующий файлы (consumer)
def worker(name: str):
    while True:
        item = asyncio.run_coroutine_threadsafe(queue_to_consume.get(), loop).result()
        logging.info(f"Начало работы над элементом с id {item['id']} by {name} "
                    f" в {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        item["worker"] = f"Worker{name}"
        item["state"] = "in_progress"
        lock = Lock()
        # блокируем словарь для выполнения не атомарных операций
        lock.acquire()
        monitoring_list[item["id"]]["state"] = "in progress"
        monitoring_list[item["id"]]["worker"] = name
        monitoring_list[item["id"]]["start_date"] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        lock.release()
        try:
            # такой код будет работать вне зависимости от ОС
            path_to_excel = Path(os.getcwd(), "files_to_send", f"{item['id']}.xlsx")
            path_to_template = Path(os.getcwd(), "Товарный-чек.xlsx")
            shutil.copyfile(path_to_template, path_to_excel)
            wb = openpyxl.load_workbook(path_to_excel)
            ws = wb.active
            ws["A1"].value = item["org_name"]
            title_cell = ws["A4"].value
            title_cell = title_cell.replace("id", str(item["id"])).replace("date", item["date"])
            ws["A4"].value = title_cell
            all_sum = 0
            i = 6
            for product in item["product_list"]:
                ws[f"A{i}"].value = i - 5
                ws[f"B{i}"].value = product[0]
                ws[f"H{i}"].value = product[1]
                ws[f"J{i}"].value = product[2]
                ws[f"M{i}"].value = product[1] * product[2]
                all_sum += product[1] * product[2]
                i += 1
            ws["M16"].value = all_sum
            wb.save(path_to_excel)
            wb.close()
            time.sleep(3)
            # блокируем словарь для выполнения не атомарных операций
            lock.acquire()
            monitoring_list[item["id"]]["state"] = "Done Successfully"
            lock.release()
        except Exception as e:
            logging.info(e)
            # блокируем словарь для выполнения не атомарных операций
            lock.acquire()
            monitoring_list[item["id"]]["state"] = "Error"
            lock.release()
        # блокируем словарь для выполнения не атомарных операций
        lock.acquire()
        monitoring_list[item["id"]]["end_date"] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        lock.release()
        queue_to_consume.task_done()
        logging.info(f"Конец работы над элементом с id {item['id']} by {name}"
                    f" в {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")


async def main():
    global loop
    loop = asyncio.get_running_loop()
    app = web.Application()
    logging.basicConfig(level=logging.DEBUG)
    logging.info("start")
    app.add_routes([web.get('/get_file', get_file)])
    app.add_routes([web.get('/queue_info', get_queue_info)])
    app.add_routes([web.post('/add_item', post_add_item)])
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, 'localhost', 8080)
    await site.start()
    loop = asyncio.get_running_loop()
    max_consumers = 2
    consumers = [Thread(target=worker, kwargs={"name": f"Worker {i}"}) for i in range(max_consumers)]
    for t in consumers:
        t.start()

    while True:
        await asyncio.sleep(3600)


if __name__ == "__main__":
    asyncio.run(main(), debug=True)
