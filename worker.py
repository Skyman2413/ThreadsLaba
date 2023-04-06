import datetime
import os
import random
import shutil
import time
from pathlib import Path
from queue import Queue
from threading import Thread, Lock

import openpyxl

queue_to_consume = Queue()
# данный словарь нужны только для мониторинга, по-хорошему он не нужен, лишний расход памяти
monitoring_list = {}


def worker_without_thread(item: {}):
    print(f"Начало работы над элементом с id {item['id']}"
          f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    item["state"] = "in_progress"
    lock = Lock()
    # блокируем словарь для выполнения не атомарных операций
    monitoring_list[item["id"]]["state"] = "in progress"
    monitoring_list[item["id"]]["start_date"] = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    try:
        path_to_excel = Path(os.getcwd(), "files_to_send", f"{item['id']}.xlsx")
        path_to_template = Path(os.getcwd(), "Товарный-чек.xlsx")
        shutil.copyfile(path_to_template, path_to_excel)
        wb = openpyxl.load_workbook(path_to_excel)
        ws = wb.active
        ws["A1"].value = item["org_name"]
        title_cell = ws["A4"].value
        title_cell = title_cell.replace("id", str(item["id"])).replace("date", item["date"])
        ws["A4"].value = title_cell
        all_sum = 0
        i = 6
        for product in item["product_list"]:
            ws[f"A{i}"].value = i - 5
            ws[f"B{i}"].value = product[0]
            ws[f"H{i}"].value = product[1]
            ws[f"J{i}"].value = product[2]
            ws[f"M{i}"].value = product[1] * product[2]
            all_sum += product[1] * product[2]
            i += 1
        ws["M16"].value = all_sum
        wb.save(path_to_excel)
        wb.close()
        time.sleep(random.randint(5, 25))
        monitoring_list[item["id"]]["state"] = "Done Successfully"
    except Exception as e:
        print(e)
        monitoring_list[item["id"]]["state"] = "Error"
    monitoring_list[item["id"]]["end_date"] = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    print(f"Конец работы над элементом с id {item['id']}"
          f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")


class Worker(Thread):
    name: str

    def __init__(self, name: str):
        super().__init__()
        self.name = name
        print(f"объект {self.name} создан")

    def run(self):
        print(f"запущен поток {self.name}")
        while True:
            item = queue_to_consume.get()
            print(f"Начало работы над элементом с id {item['id']} by {self.name} "
                  f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
            item["worker"] = f"Worker{self.name}"
            item["state"] = "in_progress"
            lock = Lock()
            # блокируем словарь для выполнения не атомарных операций
            with lock:
                monitoring_list[item["id"]]["state"] = "in progress"
                monitoring_list[item["id"]]["worker"] = self.name
                monitoring_list[item["id"]]["start_date"] = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            try:
                self.fill_excel(item)
                # блокируем словарь для выполнения не атомарных операций
                with lock:
                    monitoring_list[item["id"]]["state"] = "Done Successfully"
            except Exception as e:
                print(e)
                # блокируем словарь для выполнения не атомарных операций
                with lock:
                    monitoring_list[item["id"]]["state"] = "Error"
            # блокируем словарь для выполнения не атомарных операций
            with lock:
                monitoring_list[item["id"]]["end_date"] = datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            queue_to_consume.task_done()
            print(f"Конец работы над элементом с id {item['id']} by {self.name}"
                  f" в {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")

    def fill_excel(self, item: {}):
        # такой код будет работать вне зависимости от ОС
        path_to_excel = Path(os.getcwd(), "files_to_send", f"{item['id']}.xlsx")
        path_to_template = Path(os.getcwd(), "Товарный-чек.xlsx")
        shutil.copyfile(path_to_template, path_to_excel)
        wb = openpyxl.load_workbook(path_to_excel)
        ws = wb.active
        ws["A1"].value = item["org_name"]
        title_cell = ws["A4"].value
        title_cell = title_cell.replace("id", str(item["id"])).replace("date", item["date"])
        ws["A4"].value = title_cell
        all_sum = 0
        i = 6
        for product in item["product_list"]:
            ws[f"A{i}"].value = i - 5
            ws[f"B{i}"].value = product[0]
            ws[f"H{i}"].value = product[1]
            ws[f"J{i}"].value = product[2]
            ws[f"M{i}"].value = product[1] * product[2]
            all_sum += product[1] * product[2]
            i += 1
        ws["M16"].value = all_sum
        wb.save(path_to_excel)
        wb.close()
        time.sleep(random.randint(5, 25))
